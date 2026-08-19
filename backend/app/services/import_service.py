"""Excel/CSV journal import wizard.

Column mapping convention: each mapped field is either the name of a
column in the uploaded file, or a fixed value shared by every row,
written as "const:<value>" (e.g. "const:1010" to always post the debit
side to account 1010, which is the common case for a single bank
account's statement; a 1C-style export that already carries both sides
per row instead maps both account fields to real columns).

Required fields: entry_date, debit_account_code, credit_account_code,
amount. Optional: description, currency (defaults to the legal entity's
functional currency), cost_center, counterparty, project.

Stateless-serverless note: parsed rows are held on the ImportBatch row
itself (status="draft") between the upload/validate/commit steps, since
a Vercel function cannot rely on local disk between invocations. For
very large files this should move to chunked/streamed processing backed
by Supabase Storage instead of an in-DB JSON blob -- fine for the
register sizes this MVP targets (tens to low thousands of rows).
"""

import csv
import io
from datetime import date, datetime

from dateutil import parser as date_parser
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models.analytics import CostCenter, Counterparty, Project
from app.models.coa import ChartOfAccount
from app.models.enums import EntryDirection
from app.models.import_batch import ImportBatch
from app.models.journal import JournalEntry, JournalLine
from app.models.legal_entity import LegalEntity
from app.schemas.imports import ImportRowError
from app.services.fx_service import get_rate

REQUIRED_FIELDS = ["entry_date", "debit_account_code", "credit_account_code", "amount"]
PREVIEW_ROW_LIMIT = 20


def parse_file(file_name: str, content: bytes) -> tuple[list[str], list[dict]]:
    if file_name.lower().endswith((".xlsx", ".xlsm")):
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = [str(c) if c is not None else "" for c in next(rows_iter)]
        rows = []
        for raw in rows_iter:
            if all(c is None for c in raw):
                continue
            rows.append({header[i]: raw[i] for i in range(len(header)) if i < len(raw)})
        return header, rows

    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = [dict(r) for r in reader]
    return reader.fieldnames or [], rows


def _resolve(mapping: dict, field: str, row: dict):
    spec = mapping.get(field)
    if not spec:
        return None
    if isinstance(spec, str) and spec.startswith("const:"):
        return spec[len("const:") :]
    return row.get(spec)


def _parse_date(value) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date_parser.parse(str(value), dayfirst=True).date()


def _parse_amount(value) -> float:
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(" ", "").replace(",", ".")
    return float(s)


class ResolvedRow:
    def __init__(self):
        self.entry_date: date | None = None
        self.description = ""
        self.debit_account_id: int | None = None
        self.credit_account_id: int | None = None
        self.amount: float | None = None
        self.currency: str | None = None
        self.cost_center_id: int | None = None
        self.counterparty_id: int | None = None
        self.project_id: int | None = None


def _lookup_ci(names: dict[str, int], name: str | None) -> tuple[int | None, str | None]:
    if not name:
        return None, None
    key = str(name).strip().lower()
    if key in names:
        return names[key], None
    return None, f"unknown value '{name}'"


def resolve_and_validate(
    db: Session, legal_entity: LegalEntity, mapping: dict, rows: list[dict]
) -> tuple[list[ResolvedRow], list[ImportRowError]]:
    accounts_by_code = {
        a.code: a
        for a in db.query(ChartOfAccount).filter(ChartOfAccount.legal_entity_id == legal_entity.id)
    }
    cost_centers = {c.name.strip().lower(): c.id for c in db.query(CostCenter).filter(CostCenter.legal_entity_id == legal_entity.id)}
    counterparties = {c.name.strip().lower(): c.id for c in db.query(Counterparty).filter(Counterparty.legal_entity_id == legal_entity.id)}
    projects = {c.name.strip().lower(): c.id for c in db.query(Project).filter(Project.legal_entity_id == legal_entity.id)}

    resolved: list[ResolvedRow] = []
    errors: list[ImportRowError] = []

    missing_required = [f for f in REQUIRED_FIELDS if not mapping.get(f)]
    if missing_required:
        errors.append(ImportRowError(row_number=0, message=f"Column mapping is missing required fields: {', '.join(missing_required)}"))
        return resolved, errors

    for i, row in enumerate(rows, start=1):
        r = ResolvedRow()
        row_errors = []

        try:
            r.entry_date = _parse_date(_resolve(mapping, "entry_date", row))
        except Exception:
            row_errors.append("unparseable entry_date")

        r.description = str(_resolve(mapping, "description", row) or "")

        debit_code = _resolve(mapping, "debit_account_code", row)
        credit_code = _resolve(mapping, "credit_account_code", row)
        debit_account = accounts_by_code.get(str(debit_code)) if debit_code else None
        credit_account = accounts_by_code.get(str(credit_code)) if credit_code else None
        if debit_account is None:
            row_errors.append(f"unknown debit account code '{debit_code}'")
        elif not debit_account.is_postable:
            row_errors.append(f"account '{debit_code}' is not postable (has sub-accounts)")
        else:
            r.debit_account_id = debit_account.id
        if credit_account is None:
            row_errors.append(f"unknown credit account code '{credit_code}'")
        elif not credit_account.is_postable:
            row_errors.append(f"account '{credit_code}' is not postable (has sub-accounts)")
        else:
            r.credit_account_id = credit_account.id

        try:
            r.amount = _parse_amount(_resolve(mapping, "amount", row))
            if r.amount is not None and r.amount <= 0:
                row_errors.append("amount must be positive")
        except Exception:
            row_errors.append("unparseable amount")

        r.currency = str(_resolve(mapping, "currency", row) or legal_entity.functional_currency).upper()

        cc_id, cc_err = _lookup_ci(cost_centers, _resolve(mapping, "cost_center", row))
        if cc_err:
            row_errors.append(f"cost center: {cc_err}")
        r.cost_center_id = cc_id

        cp_id, cp_err = _lookup_ci(counterparties, _resolve(mapping, "counterparty", row))
        if cp_err:
            row_errors.append(f"counterparty: {cp_err}")
        r.counterparty_id = cp_id

        pj_id, pj_err = _lookup_ci(projects, _resolve(mapping, "project", row))
        if pj_err:
            row_errors.append(f"project: {pj_err}")
        r.project_id = pj_id

        if row_errors:
            errors.append(ImportRowError(row_number=i, message="; ".join(row_errors)))
        resolved.append(r)

    return resolved, errors


def commit_rows(
    db: Session,
    legal_entity: LegalEntity,
    resolved_rows: list[ResolvedRow],
    row_errors: list[ImportRowError],
    import_batch: ImportBatch,
    created_by_user_id: int | None,
) -> tuple[int, int]:
    bad_row_numbers = {e.row_number for e in row_errors}
    entries_created = 0
    lines_created = 0

    for i, r in enumerate(resolved_rows, start=1):
        if i in bad_row_numbers:
            continue

        fx_to_local = get_rate(db, r.entry_date, r.currency, legal_entity.functional_currency)
        fx_to_usd = get_rate(db, r.entry_date, r.currency, "USD")
        local_amount = round(r.amount * fx_to_local, 2)
        usd_amount = round(r.amount * fx_to_usd, 2)

        entry = JournalEntry(
            legal_entity_id=legal_entity.id,
            entry_date=r.entry_date,
            description=r.description,
            created_by_user_id=created_by_user_id,
            import_batch_id=import_batch.id,
        )
        for account_id, direction in ((r.debit_account_id, EntryDirection.DEBIT), (r.credit_account_id, EntryDirection.CREDIT)):
            entry.lines.append(
                JournalLine(
                    account_id=account_id,
                    cost_center_id=r.cost_center_id,
                    counterparty_id=r.counterparty_id,
                    project_id=r.project_id,
                    direction=direction,
                    transaction_currency=r.currency,
                    transaction_amount=r.amount,
                    local_currency_amount=local_amount,
                    usd_amount=usd_amount,
                    fx_rate_to_local=fx_to_local,
                    fx_rate_to_usd=fx_to_usd,
                    memo=r.description,
                )
            )
        db.add(entry)
        entries_created += 1
        lines_created += 2

    db.flush()
    return entries_created, lines_created
