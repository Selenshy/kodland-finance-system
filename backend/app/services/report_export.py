import io

from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet

from app.schemas.reports import ReportOut

REPORT_TITLES = {"pl": "Profit & Loss", "cf": "Cash Flow (Direct Method)", "balance": "Balance Sheet"}


def to_excel(report: ReportOut) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = REPORT_TITLES.get(report.report_type, report.report_type)[:31]

    bold = Font(bold=True)
    ws.append([REPORT_TITLES.get(report.report_type, report.report_type)])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Period: {report.period_start} - {report.period_end}", f"Currency: {report.currency}"])
    ws.append([])

    for section in report.sections:
        ws.append([section.title])
        ws.cell(row=ws.max_row, column=1).font = bold
        for line in section.lines:
            ws.append([line.label, line.amount])
            if line.is_subtotal:
                ws.cell(row=ws.max_row, column=1).font = bold
                ws.cell(row=ws.max_row, column=2).font = bold
        ws.append(["Total", section.total])
        ws.cell(row=ws.max_row, column=1).font = bold
        ws.cell(row=ws.max_row, column=2).font = bold
        ws.append([])

    if report.check_ok is not None:
        ws.append(["Assets = Liabilities + Equity check", "OK" if report.check_ok else "MISMATCH"])

    for col, width in (("A", 40), ("B", 20)):
        ws.column_dimensions[col].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_pdf(report: ReportOut) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    story = [
        Paragraph(REPORT_TITLES.get(report.report_type, report.report_type), styles["Title"]),
        Paragraph(f"Period: {report.period_start} to {report.period_end} &nbsp;|&nbsp; Currency: {report.currency}", styles["Normal"]),
        Spacer(1, 0.5 * cm),
    ]

    for section in report.sections:
        story.append(Paragraph(section.title, styles["Heading2"]))
        data = [["Line", "Amount"]] + [[line.label, f"{line.amount:,.2f}"] for line in section.lines]
        data.append(["Total", f"{section.total:,.2f}"])
        table = Table(data, colWidths=[10 * cm, 5 * cm])
        table.setStyle(
            TableStyle(
                [
                    ("FONT", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONT", (0, -1), (-1, -1), "Helvetica-Bold"),
                    ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.grey),
                    ("LINEABOVE", (0, -1), (-1, -1), 0.5, colors.grey),
                    ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ]
            )
        )
        story.append(table)
        story.append(Spacer(1, 0.5 * cm))

    if report.check_ok is not None:
        story.append(
            Paragraph(
                f"Assets = Liabilities + Equity check: {'OK' if report.check_ok else 'MISMATCH'}",
                styles["Normal"],
            )
        )

    doc.build(story)
    return buf.getvalue()
